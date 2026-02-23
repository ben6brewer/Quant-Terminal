"""Portfolio Export Service - CSV and Excel export logic."""

import csv
from typing import List, Dict, Any, Tuple


class PortfolioExportService:
    """Handles data preparation and file export for portfolio data."""

    @staticmethod
    def get_transaction_export_data(
        transactions: List[Dict[str, Any]],
        cached_prices: Dict[str, float],
        cached_names: Dict[str, str],
        historical_prices: Dict[str, Dict[str, float]],
        portfolio_name: str,
    ) -> Tuple[List[Dict[str, Any]], List[str], str]:
        """
        Build transaction rows for export.

        Returns:
            Tuple of (data rows, column headers, filename prefix)
        """
        columns = [
            "Date", "Ticker", "Name", "Quantity", "Execution Price",
            "Fees", "Type", "Daily Closing Price", "Live Price",
            "Principal", "Market Value"
        ]

        export_data = []
        for tx in transactions:
            ticker = tx.get("ticker", "")
            quantity = tx.get("quantity", 0) or 0
            entry_price = tx.get("entry_price", 0) or 0
            fees = tx.get("fees", 0) or 0
            current_price = cached_prices.get(ticker, 0) or 0

            tx_date = tx.get("date", "")
            daily_close = historical_prices.get(ticker, {}).get(tx_date, "")

            principal = quantity * entry_price + fees
            market_value = quantity * current_price if current_price else ""

            row = {
                "Date": tx_date,
                "Ticker": ticker,
                "Name": cached_names.get(ticker, ""),
                "Quantity": quantity,
                "Execution Price": entry_price,
                "Fees": fees,
                "Type": tx.get("transaction_type", ""),
                "Daily Closing Price": daily_close,
                "Live Price": current_price if current_price else "",
                "Principal": principal,
                "Market Value": market_value
            }
            export_data.append(row)

        return export_data, columns, f"{portfolio_name}_transactions"

    @staticmethod
    def get_holdings_export_data(
        holdings_data: List[Dict[str, Any]],
        cached_names: Dict[str, str],
        portfolio_name: str,
    ) -> Tuple[List[Dict[str, Any]], List[str], str]:
        """
        Build holdings rows for export (excludes TOTAL row).

        Returns:
            Tuple of (data rows, column headers, filename prefix)
        """
        columns = [
            "Ticker", "Name", "Quantity", "Avg Cost Basis",
            "Current Price", "Market Value", "P&L", "Weight %"
        ]

        export_data = []
        for holding in holdings_data:
            ticker = holding.get("ticker", "")
            is_free_cash = holding.get("_is_free_cash", False)

            row = {
                "Ticker": ticker,
                "Name": cached_names.get(ticker, "") if not is_free_cash else "",
                "Quantity": holding.get("total_quantity", 0),
                "Avg Cost Basis": holding.get("avg_cost_basis", 0),
                "Current Price": holding.get("current_price", "") if not is_free_cash else "",
                "Market Value": holding.get("market_value", ""),
                "P&L": holding.get("total_pnl", 0) if not is_free_cash else "",
                "Weight %": holding.get("weight_pct", 0)
            }
            export_data.append(row)

        return export_data, columns, f"{portfolio_name}_holdings"

    @staticmethod
    def export_to_csv(
        parent,
        theme_manager,
        data: List[Dict[str, Any]],
        columns: List[str],
        filename_prefix: str,
    ):
        """Export data to CSV file with a save dialog."""
        from PySide6.QtWidgets import QFileDialog
        from app.ui.widgets.common import CustomMessageBox

        default_name = f"{filename_prefix}.csv"
        file_path, _ = QFileDialog.getSaveFileName(
            parent,
            "Export to CSV",
            default_name,
            "CSV Files (*.csv);;All Files (*)"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=columns)
                writer.writeheader()
                writer.writerows(data)

            CustomMessageBox.information(
                theme_manager,
                parent,
                "Export Complete",
                f"Data exported to:\n{file_path}"
            )
        except Exception as e:
            CustomMessageBox.critical(
                theme_manager,
                parent,
                "Export Error",
                f"Failed to export data:\n{str(e)}"
            )

    @staticmethod
    def export_to_excel(
        parent,
        theme_manager,
        data: List[Dict[str, Any]],
        columns: List[str],
        sheet_name: str,
    ):
        """
        Export data to Excel.

        On Windows: Uses COM automation to open Excel directly.
        On macOS/Linux: Creates .xlsx file with openpyxl and opens with default app.
        """
        import sys

        if sys.platform == 'win32':
            PortfolioExportService._export_to_excel_windows(
                parent, theme_manager, data, columns, sheet_name
            )
        else:
            PortfolioExportService._export_to_excel_crossplatform(
                parent, theme_manager, data, columns, sheet_name
            )

    @staticmethod
    def _export_to_excel_windows(
        parent,
        theme_manager,
        data: List[Dict[str, Any]],
        columns: List[str],
        sheet_name: str,
    ):
        """Export to Excel using Windows COM automation."""
        from app.ui.widgets.common import CustomMessageBox

        try:
            import win32com.client
        except ImportError:
            PortfolioExportService._export_to_excel_crossplatform(
                parent, theme_manager, data, columns, sheet_name
            )
            return

        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = True

            workbook = excel.Workbooks.Add()
            worksheet = workbook.ActiveSheet
            worksheet.Name = sheet_name[:31]

            for col_idx, col_name in enumerate(columns, start=1):
                worksheet.Cells(1, col_idx).Value = col_name
                worksheet.Cells(1, col_idx).Font.Bold = True

            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    value = row_data.get(col_name, "")
                    worksheet.Cells(row_idx, col_idx).Value = value

            worksheet.Columns.AutoFit()

        except Exception as e:
            CustomMessageBox.critical(
                theme_manager,
                parent,
                "Excel Error",
                f"Failed to open Excel:\n{str(e)}\n\n"
                "Make sure Microsoft Excel is installed."
            )

    @staticmethod
    def _export_to_excel_crossplatform(
        parent,
        theme_manager,
        data: List[Dict[str, Any]],
        columns: List[str],
        sheet_name: str,
    ):
        """Export to Excel using openpyxl (works on macOS/Linux)."""
        import subprocess
        import sys
        import tempfile
        import os
        from app.ui.widgets.common import CustomMessageBox

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            CustomMessageBox.warning(
                theme_manager,
                parent,
                "Excel Not Available",
                "Excel export requires the 'openpyxl' package.\n\n"
                "Install it with: pip install openpyxl"
            )
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]

            bold_font = Font(bold=True)
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = bold_font

            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    value = row_data.get(col_name, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            for col_idx, col_name in enumerate(columns, start=1):
                max_length = len(str(col_name))
                for row_data in data:
                    cell_value = str(row_data.get(col_name, ""))
                    max_length = max(max_length, len(cell_value))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

            temp_dir = tempfile.gettempdir()
            file_path = os.path.join(temp_dir, f"{sheet_name}.xlsx")
            wb.save(file_path)

            if sys.platform == 'darwin':
                subprocess.run(['open', file_path], check=True)
            elif sys.platform == 'win32':
                os.startfile(file_path)
            else:
                subprocess.run(['xdg-open', file_path], check=True)

        except Exception as e:
            CustomMessageBox.critical(
                theme_manager,
                parent,
                "Excel Error",
                f"Failed to create Excel file:\n{str(e)}"
            )
