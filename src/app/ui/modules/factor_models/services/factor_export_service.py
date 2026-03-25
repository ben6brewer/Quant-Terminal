"""Factor Export Service — Excel export for factor model regression results."""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import pandas as pd
    from .factor_regression_service import FactorRegressionResult
    from .model_definitions import FactorModelSpec

logger = logging.getLogger(__name__)

# ── Source metadata ──────────────────────────────────────────────────────────

_FF_FACTORS = {"Mkt-RF", "SMB", "HML", "RMW", "CMA", "UMD"}
_Q_FACTORS = {"R_MKT", "R_ME", "R_IA", "R_ROE", "R_EG"}
_AQR_FACTORS = {"BAB", "QMJ", "HML_DEVIL"}

_SOURCE_INFO = {
    "ff": (
        "Fama-French",
        "Kenneth French Data Library \u2014 Market, Size, Value, Profitability, Investment, Momentum factors",
        "https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/data_library.html",
    ),
    "q": (
        "Q-Factor (HXZ)",
        "Hou-Xue-Zhang Q-factor model \u2014 Market, Size, Investment, Profitability, Expected Growth",
        "http://global-q.org",
    ),
    "aqr": (
        "AQR",
        "AQR Capital Management \u2014 Betting Against Beta, Quality Minus Junk, HML Devil",
        "https://www.aqr.com/Insights/Datasets",
    ),
}


class FactorExportService:
    """Export factor regression results to a formatted Excel workbook."""

    @staticmethod
    def export_to_excel(
        parent,
        theme_manager,
        result: "FactorRegressionResult",
        factor_returns_df: "pd.DataFrame",
        rf_series: "pd.Series",
        model_spec: "FactorModelSpec",
        identifier: str = "",
    ) -> None:
        """Open a save dialog and write a 3-sheet Excel workbook.

        Sheets:
            1. Factor Returns — raw factor return time series
            2. Regression Results — coefficients, goodness of fit, diagnostics
            3. Sources — data source descriptions and URLs
        """
        from PySide6.QtWidgets import QFileDialog
        from app.ui.widgets.common import CustomMessageBox

        # Default filename
        file_label = identifier or model_spec.name.replace(" ", "_")
        default_name = f"factor_regression_{file_label}_{result.frequency}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            parent,
            "Export Factor Regression",
            default_name,
            "Excel Files (*.xlsx);;All Files (*)",
        )
        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, numbers

            wb = Workbook()

            # ── Sheet 1: Factor Returns ──────────────────────────────────
            ws1 = wb.active
            ws1.title = "Factor Returns"
            ws1.sheet_properties.tabColor = "4472C4"

            _write_factor_returns(ws1, factor_returns_df, rf_series, result, identifier)

            # ── Sheet 2: Regression Results ──────────────────────────────
            ws2 = wb.create_sheet("Regression Results")
            ws2.sheet_properties.tabColor = "70AD47"

            _write_regression_results(ws2, result)

            # ── Sheet 3: Sources ─────────────────────────────────────────
            ws3 = wb.create_sheet("Sources")
            ws3.sheet_properties.tabColor = "ED7D31"

            _write_sources(ws3, result, model_spec)

            wb.save(file_path)

            CustomMessageBox.information(
                theme_manager,
                parent,
                "Export Complete",
                f"Regression results exported to:\n{file_path}",
            )

        except ImportError:
            CustomMessageBox.warning(
                theme_manager,
                parent,
                "Excel Not Available",
                "Excel export requires the 'openpyxl' package.\n\n"
                "Install it with: pip install openpyxl",
            )
        except Exception as e:
            logger.exception("Factor export failed")
            CustomMessageBox.critical(
                theme_manager,
                parent,
                "Export Error",
                f"Failed to export:\n{e}",
            )


# ── Sheet writers ────────────────────────────────────────────────────────────

_HEADER_FONT = None
_HEADER_FILL = None
_SECTION_FONT = None
_NUM_4 = "0.0000"
_NUM_6 = "0.000000"
_NUM_2 = "0.00"
_NUM_INT = "0"


def _styles():
    """Lazily create openpyxl style objects."""
    global _HEADER_FONT, _HEADER_FILL, _SECTION_FONT
    if _HEADER_FONT is None:
        from openpyxl.styles import Font, PatternFill

        _HEADER_FONT = Font(bold=True, size=11)
        _HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
        _SECTION_FONT = Font(bold=True, size=12)
    return _HEADER_FONT, _HEADER_FILL, _SECTION_FONT


def _auto_width(ws, min_width: int = 10, max_width: int = 30):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        lengths = []
        for cell in col:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        if lengths:
            best = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[col[0].column_letter].width = best


def _write_header_row(ws, row: int, values: list[str]):
    """Write a bold header row with fill."""
    hf, hfill, _ = _styles()
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = hf
        cell.fill = hfill


def _write_factor_returns(ws, factor_df: "pd.DataFrame", rf_series, result, identifier: str = ""):
    """Sheet 1 — asset excess returns + risk-free rate + raw factor return time series."""
    import pandas as pd
    import numpy as np

    # Use the dates from the regression result to align
    dates = pd.to_datetime(result.dates)
    factor_cols = list(result.factor_names)

    # Slice factor_df to regression dates
    common_idx = factor_df.index.intersection(dates)
    df = factor_df.loc[common_idx, [c for c in factor_cols if c in factor_df.columns]]

    # Build asset excess returns series aligned to factor dates
    asset_label = identifier.upper() if identifier else "Asset"
    asset_series = pd.Series(result.asset_excess_returns, index=dates, name=asset_label)

    # Align risk-free rate to the same dates
    rf_aligned = rf_series.reindex(common_idx)

    # Header: Date | Ticker returns | US10Y (RF) | Factor columns...
    headers = ["Date", asset_label, "US10Y (RF)"] + list(df.columns)
    _write_header_row(ws, 1, headers)

    # Data
    for row_idx, (dt, row_data) in enumerate(df.iterrows(), start=2):
        ws.cell(row=row_idx, column=1, value=dt.strftime("%Y-%m-%d"))
        # Column B: asset excess return
        asset_val = asset_series.get(dt, np.nan)
        if not np.isnan(asset_val):
            ws.cell(row=row_idx, column=2, value=float(asset_val)).number_format = _NUM_6
        # Column C: risk-free rate
        rf_val = rf_aligned.get(dt, np.nan)
        if not (isinstance(rf_val, float) and np.isnan(rf_val)):
            ws.cell(row=row_idx, column=3, value=float(rf_val)).number_format = _NUM_6
        # Columns D+: factor returns
        for col_idx, col_name in enumerate(df.columns, start=4):
            cell = ws.cell(row=row_idx, column=col_idx, value=float(row_data[col_name]))
            cell.number_format = _NUM_6

    _auto_width(ws, min_width=12)


def _write_regression_results(ws, result):
    """Sheet 2 — coefficients, goodness of fit, diagnostics."""
    _, _, section_font = _styles()
    row = 1

    # ── Section A: Coefficients ──────────────────────────────────────
    ws.cell(row=row, column=1, value="Coefficients").font = section_font
    row += 1

    coeff_headers = [
        "Parameter", "Coefficient", "Std Error",
        "T-Statistic", "P-Value", "95% CI Lower", "95% CI Upper",
    ]
    _write_header_row(ws, row, coeff_headers)
    row += 1

    # Alpha row
    row = _write_coeff_row(ws, row, "Alpha (annualized)", result.alpha_annualized, result)
    # Factor rows
    for fname in result.factor_names:
        beta = result.betas.get(fname, 0.0)
        row = _write_coeff_row(ws, row, fname, beta, result)

    row += 2

    # ── Section B: Goodness of Fit ───────────────────────────────────
    ws.cell(row=row, column=1, value="Goodness of Fit").font = section_font
    row += 1

    fit_headers = ["Metric", "Value"]
    _write_header_row(ws, row, fit_headers)
    row += 1

    fit_data = [
        ("R-Squared", result.r_squared, _NUM_4),
        ("Adj R-Squared", result.adj_r_squared, _NUM_4),
        ("F-Statistic", result.f_statistic, _NUM_2),
        ("Prob(F-Statistic)", result.f_p_value, _NUM_6),
    ]
    for label, value, fmt in fit_data:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = fmt
        row += 1

    row += 2

    # ── Section C: Diagnostics ───────────────────────────────────────
    ws.cell(row=row, column=1, value="Diagnostics").font = section_font
    row += 1

    _write_header_row(ws, row, ["Metric", "Value"])
    row += 1

    diag_data = [
        ("Durbin-Watson", result.durbin_watson, _NUM_4),
        ("Residual Std Error", result.residual_std_error, _NUM_6),
        ("N Observations", result.n_observations, _NUM_INT),
        ("Frequency", result.frequency, None),
        ("Annualization Factor", result.annualization_factor, _NUM_INT),
    ]
    for label, value, fmt in diag_data:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        if fmt:
            cell.number_format = fmt
        row += 1

    _auto_width(ws, min_width=14)


def _write_coeff_row(ws, row: int, label: str, coeff: float, result) -> int:
    """Write a single coefficient row and return next row number."""
    key = "Alpha" if "Alpha" in label else label
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=coeff).number_format = _NUM_4
    ws.cell(row=row, column=3, value=result.std_errors.get(key, 0.0)).number_format = _NUM_4
    ws.cell(row=row, column=4, value=result.t_stats.get(key, 0.0)).number_format = _NUM_4
    ws.cell(row=row, column=5, value=result.p_values.get(key, 0.0)).number_format = _NUM_6
    ws.cell(row=row, column=6, value=result.ci_lower.get(key, 0.0)).number_format = _NUM_4
    ws.cell(row=row, column=7, value=result.ci_upper.get(key, 0.0)).number_format = _NUM_4
    return row + 1


def _write_sources(ws, result, model_spec):
    """Sheet 3 — model metadata and data source links."""
    import pandas as pd

    _, _, section_font = _styles()

    # Metadata header
    ws.cell(row=1, column=1, value="Model:").font = _styles()[0]
    ws.cell(row=1, column=2, value=model_spec.name)

    ws.cell(row=2, column=1, value="Frequency:").font = _styles()[0]
    ws.cell(row=2, column=2, value=result.frequency.capitalize())

    dates = pd.to_datetime(result.dates)
    if len(dates) > 0:
        ws.cell(row=3, column=1, value="Date Range:").font = _styles()[0]
        ws.cell(
            row=3, column=2,
            value=f"{dates[0].strftime('%Y-%m-%d')} to {dates[-1].strftime('%Y-%m-%d')}",
        )

    ws.cell(row=4, column=1, value="Factors:").font = _styles()[0]
    ws.cell(row=4, column=2, value=", ".join(result.factor_names))

    row = 6

    # Source table
    ws.cell(row=row, column=1, value="Data Sources").font = section_font
    row += 1

    _write_header_row(ws, row, ["Source", "Description", "URL"])
    row += 1

    # Determine which sources are used
    factors_set = set(model_spec.factors)
    used_sources = []
    if factors_set & _FF_FACTORS:
        used_sources.append("ff")
    if factors_set & _Q_FACTORS:
        used_sources.append("q")
    if factors_set & _AQR_FACTORS:
        used_sources.append("aqr")

    for src_key in used_sources:
        name, desc, url = _SOURCE_INFO[src_key]
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=url)
        row += 1

    _auto_width(ws, min_width=12, max_width=80)
